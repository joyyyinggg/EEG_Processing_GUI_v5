# -*- coding: utf-8 -*-
"""ecg/exporter.py — Excel export for ECG results."""
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def export_excel(path: str, results: list, hr_high: float, hrv_low: float):
    wb = Workbook(); ws = wb.active; ws.title = 'Heart Rate Results'
    thin   = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill('solid', fgColor='1D4ED8')
    h_font = Font(bold=True, color='FFFFFF', size=12)
    headers = ['Filename','HR (BPM)','HRV RMSSD (ms)','R-peaks','Status','Note']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    fill_map = {
        'Calm':         PatternFill('solid', fgColor='DCFCE7'),
        'Stressed':     PatternFill('solid', fgColor='FECDD3'),
        'Bradycardia':  PatternFill('solid', fgColor='FEF3C7'),
        'Poor quality': PatternFill('solid', fgColor='FFF3CD'),
    }
    for row_i, rec in enumerate(results, 2):
        vals = [rec['Filename'], rec['HR (BPM)'], rec['HRV RMSSD (ms)'],
                rec['R-peaks'], rec['Status'], rec.get('Note','')]
        fill = fill_map.get(rec['Status'], PatternFill('solid', fgColor='FFFFFF'))
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col == 5: cell.font = Font(bold=True, size=12)
    for col, w in enumerate([30,12,18,12,14,30], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws2 = wb.create_sheet('Summary')
    hrs = [r['HR (BPM)'] for r in results if isinstance(r['HR (BPM)'], (int,float))]
    stats = [
        ('Total files', len(results)),
        ('Calm',        sum(1 for r in results if r['Status']=='Calm')),
        ('Stressed',    sum(1 for r in results if r['Status']=='Stressed')),
        ('Bradycardia', sum(1 for r in results if r['Status']=='Bradycardia')),
        ('Poor quality',sum(1 for r in results if r['Status']=='Poor quality')),
        ('Mean HR (BPM)', round(float(np.mean(hrs)),1) if hrs else 'N/A'),
        ('SD HR (BPM)',   round(float(np.std(hrs)),1)  if hrs else 'N/A'),
        ('Stressed threshold', f'HR > {hr_high} BPM'),
        ('HRV threshold',      f'RMSSD < {hrv_low} ms'),
    ]
    for r,(label,val) in enumerate(stats,1):
        ws2.cell(r,1,label).font = Font(bold=True, size=12)
        ws2.cell(r,2,val).font   = Font(size=12)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18
    wb.save(path)
